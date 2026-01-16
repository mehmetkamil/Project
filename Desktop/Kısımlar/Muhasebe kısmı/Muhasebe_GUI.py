#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🏢 AXA Muhasebe ve Komisyon Takip Sistemi - GUI
================================================
Özellikler:
- Kişi bazlı tab yapısı (YAŞAR, KAMİL, TEZER)
- PDF otomatik okuma ve veri çıkarma
- Otomatik komisyon hesaplama
- Aylık/yıllık filtreleme
- Excel raporlama
- Modern ve kullanıcı dostu arayüz
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import sys
import sqlite3
import pandas as pd
from datetime import datetime
from pathlib import Path

# Çoklu şirket parser modülünü import et
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from multi_parser import (
        process_pdf,
        process_axa_pdf,
        hesapla_komisyon,
        parse_turkish_amount
    )
    PARSER_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ Uyarı: multi_parser.py modülü bulunamadı: {e}")
    PARSER_AVAILABLE = False

# ==============================================================================
# 🎨 RENKLER VE SABITLER
# ==============================================================================

COLORS = {
    'primary': '#2563eb',
    'success': '#16a34a',
    'warning': '#ea580c',
    'danger': '#dc2626',
    'info': '#0891b2',
    'bg_light': '#f8fafc',
    'text_dark': '#1e293b',
    'text_light': '#ffffff'
}

AYLAR = ['OCAK', 'ŞUBAT', 'MART', 'NİSAN', 'MAYIS', 'HAZİRAN', 
         'TEMMUZ', 'AĞUSTOS', 'EYLÜL', 'EKİM', 'KASIM', 'ARALIK']

KISILER = ['YAŞAR', 'KAMİL', 'TEZER', 'CMC']

TURLER = ['TRAFİK', 'KASKO', 'DASK', 'EVİM', 'SAĞLIK', 'SEYAHAT', 'NAKLİYAT', 'İŞYERİ']

ODEME_TURLERI = ['AÇIK', 'K.KART']

# ==============================================================================
# 🗄️ VERİTABANI YÖNETİMİ
# ==============================================================================

class MuhasebeDB:
    """Muhasebe veritabanı yönetimi sınıfı"""
    
    def __init__(self, db_path=None):
        if db_path is None:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            db_path = os.path.join(script_dir, 'muhasebe.db')
        self.db_path = db_path
        self.init_database()
    
    def init_database(self):
        """Veritabanını oluştur - AXA ve Diğer Şirketler için ayrı tablolar"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # AXA Poliçeleri Tablosu
                cursor.execute('''
                CREATE TABLE IF NOT EXISTS komisyonlar_axa (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    kisi TEXT NOT NULL,
                    ay TEXT NOT NULL,
                    yil INTEGER NOT NULL,
                    sigortali TEXT,
                    police_no TEXT UNIQUE,
                    tarih TEXT,
                    plaka TEXT,
                    tur TEXT,
                    odeme_turu TEXT DEFAULT 'AÇIK',
                    kart_sahibi TEXT,
                    brut_prim REAL DEFAULT 0,
                    tramer REAL DEFAULT 0,
                    net_prim REAL DEFAULT 0,
                    komisyon_orani REAL DEFAULT 0,
                    toplam_komisyon REAL DEFAULT 0,
                    odeme_orani REAL DEFAULT 0,
                    odenen_komisyon REAL DEFAULT 0,
                    ikinci_police INTEGER DEFAULT 0,
                    notlar TEXT,
                    kayit_tarihi DATETIME DEFAULT CURRENT_TIMESTAMP
                )
                ''')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_kisi_axa ON komisyonlar_axa(kisi)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_ay_yil_axa ON komisyonlar_axa(ay, yil)')
                
                # Diğer Şirketler Tablosu
                cursor.execute('''
                CREATE TABLE IF NOT EXISTS komisyonlar_other (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    kisi TEXT NOT NULL,
                    ay TEXT NOT NULL,
                    yil INTEGER NOT NULL,
                    sigortali TEXT,
                    police_no TEXT UNIQUE,
                    tarih TEXT,
                    plaka TEXT,
                    tur TEXT,
                    odeme_turu TEXT DEFAULT 'AÇIK',
                    kart_sahibi TEXT,
                    brut_prim REAL DEFAULT 0,
                    tramer REAL DEFAULT 0,
                    net_prim REAL DEFAULT 0,
                    komisyon_orani REAL DEFAULT 0,
                    toplam_komisyon REAL DEFAULT 0,
                    odeme_orani REAL DEFAULT 0,
                    odenen_komisyon REAL DEFAULT 0,
                    ikinci_police INTEGER DEFAULT 0,
                    sirket TEXT NOT NULL,
                    notlar TEXT,
                    kayit_tarihi DATETIME DEFAULT CURRENT_TIMESTAMP
                )
                ''')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_kisi_other ON komisyonlar_other(kisi)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_ay_yil_other ON komisyonlar_other(ay, yil)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_sirket ON komisyonlar_other(sirket)')
                
                # Eski 'komisyonlar' tablosundan veri migration
                try:
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='komisyonlar'")
                    if cursor.fetchone():
                        # AXA kayıtlarını axa tablosuna taşı (sirket sütunu olmadan)
                        cursor.execute("""
                            INSERT OR IGNORE INTO komisyonlar_axa 
                            (id, kisi, ay, yil, sigortali, police_no, tarih, plaka, tur, odeme_turu, 
                             kart_sahibi, brut_prim, tramer, net_prim, komisyon_orani, toplam_komisyon, 
                             odeme_orani, odenen_komisyon, ikinci_police, notlar, kayit_tarihi)
                            SELECT id, kisi, ay, yil, sigortali, police_no, tarih, plaka, tur, odeme_turu, 
                                   kart_sahibi, brut_prim, COALESCE(tramer, 0), net_prim, komisyon_orani, toplam_komisyon, 
                                   odeme_orani, odenen_komisyon, COALESCE(ikinci_police, 0), notlar, kayit_tarihi
                            FROM komisyonlar WHERE sirket='AXA' OR sirket IS NULL
                        """)
                        # Diğer kayıtları other tablosuna taşı
                        cursor.execute("""
                            INSERT OR IGNORE INTO komisyonlar_other 
                            (id, kisi, ay, yil, sigortali, police_no, tarih, plaka, tur, odeme_turu, 
                             kart_sahibi, brut_prim, tramer, net_prim, komisyon_orani, toplam_komisyon, 
                             odeme_orani, odenen_komisyon, ikinci_police, sirket, notlar, kayit_tarihi)
                            SELECT id, kisi, ay, yil, sigortali, police_no, tarih, plaka, tur, odeme_turu, 
                                   kart_sahibi, brut_prim, COALESCE(tramer, 0), net_prim, komisyon_orani, toplam_komisyon, 
                                   odeme_orani, odenen_komisyon, COALESCE(ikinci_police, 0), sirket, notlar, kayit_tarihi
                            FROM komisyonlar WHERE sirket IS NOT NULL AND sirket != 'AXA'
                        """)
                        # Eski tabloyu yedekle
                        cursor.execute("ALTER TABLE komisyonlar RENAME TO komisyonlar_backup")
                except Exception as e:
                    print(f"Migration hatası (normal olabilir): {e}")
                
                conn.commit()
            return True
        except Exception as e:
            print(f"Veritabanı hatası: {e}")
            return False
    
    def insert_record(self, data):
        """Yeni kayıt ekle - şirkete göre doğru tabloya"""
        try:
            sirket = data.get('sirket', 'AXA').upper()
            table_name = 'komisyonlar_axa' if sirket == 'AXA' else 'komisyonlar_other'
            
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                if table_name == 'komisyonlar_axa':
                    cursor.execute('''
                    INSERT INTO komisyonlar_axa 
                    (kisi, ay, yil, sigortali, police_no, tarih, plaka, tur, odeme_turu, 
                     kart_sahibi, brut_prim, tramer, net_prim, komisyon_orani, toplam_komisyon, 
                     odeme_orani, odenen_komisyon, ikinci_police, notlar)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        data.get('kisi', ''),
                        data.get('ay', ''),
                        data.get('yil', datetime.now().year),
                        data.get('sigortali', ''),
                        data.get('police_no', ''),
                        data.get('tarih', ''),
                        data.get('plaka', '-'),
                        data.get('tur', ''),
                        data.get('odeme_turu', 'AÇIK'),
                        data.get('kart_sahibi', ''),
                        data.get('brut_prim', 0),
                        data.get('tramer', 0),
                        data.get('net_prim', 0),
                        data.get('komisyon_orani', 0),
                        data.get('toplam_komisyon', 0),
                        data.get('odeme_orani', 0),
                        data.get('odenen_komisyon', 0),
                        data.get('ikinci_police', 0),
                        data.get('notlar', '')
                    ))
                else:
                    cursor.execute('''
                    INSERT INTO komisyonlar_other 
                    (kisi, ay, yil, sigortali, police_no, tarih, plaka, tur, odeme_turu, 
                     kart_sahibi, brut_prim, tramer, net_prim, komisyon_orani, toplam_komisyon, 
                     odeme_orani, odenen_komisyon, ikinci_police, sirket, notlar)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        data.get('kisi', ''),
                        data.get('ay', ''),
                        data.get('yil', datetime.now().year),
                        data.get('sigortali', ''),
                        data.get('police_no', ''),
                        data.get('tarih', ''),
                        data.get('plaka', '-'),
                        data.get('tur', ''),
                        data.get('odeme_turu', 'AÇIK'),
                        data.get('kart_sahibi', ''),
                        data.get('brut_prim', 0),
                        data.get('tramer', 0),
                        data.get('net_prim', 0),
                        data.get('komisyon_orani', 0),
                        data.get('toplam_komisyon', 0),
                        data.get('odeme_orani', 0),
                        data.get('odenen_komisyon', 0),
                        data.get('ikinci_police', 0),
                        sirket,
                        data.get('notlar', '')
                    ))
                
                conn.commit()
                return True
        except sqlite3.IntegrityError:
            return False  # Duplicate
        except Exception as e:
            print(f"Kayıt ekleme hatası: {e}")
            return False
    
    def get_records(self, kisi=None, yil=None):
        """Kayıtları getir - her iki tablodan"""
        try:
            conn = sqlite3.connect(self.db_path)
            
            # AXA kayıtları
            query_axa = "SELECT *, 'AXA' as sirket FROM komisyonlar_axa WHERE 1=1"
            params = []
            
            if kisi:
                query_axa += " AND kisi = ?"
                params.append(kisi)
            if yil:
                query_axa += " AND yil = ?"
                params.append(yil)
            
            df_axa = pd.read_sql_query(query_axa, conn, params=params)
            
            # Diğer şirket kayıtları
            query_other = "SELECT * FROM komisyonlar_other WHERE 1=1"
            params_other = []
            
            if kisi:
                query_other += " AND kisi = ?"
                params_other.append(kisi)
            if yil:
                query_other += " AND yil = ?"
                params_other.append(yil)
            
            df_other = pd.read_sql_query(query_other, conn, params=params_other)
            
            # Birleştir ve sırala
            df = pd.concat([df_axa, df_other], ignore_index=True)
            df = df.sort_values('id', ascending=False)
            
            conn.close()
            return df
        except Exception as e:
            print(f"Kayıt getirme hatası: {e}")
            return pd.DataFrame()
    
    def delete_record(self, record_id, table_type='axa'):
        """Kayıt sil - hangi tablodan olduğunu belirt"""
        try:
            table_name = 'komisyonlar_axa' if table_type == 'axa' else 'komisyonlar_other'
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute(f"DELETE FROM {table_name} WHERE id = ?", (record_id,))
                conn.commit()
                return True
        except Exception as e:
            print(f"Silme hatası: {e}")
            return False
    
    def update_record(self, record_id, table_type='axa', **fields):
        """Kayıt güncelle - hangi tabloda olduğunu belirt"""
        try:
            table_name = 'komisyonlar_axa' if table_type == 'axa' else 'komisyonlar_other'
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                updates = []
                params = []
                for key, value in fields.items():
                    updates.append(f"{key} = ?")
                    params.append(value)
                params.append(record_id)
                query = f"UPDATE {table_name} SET {', '.join(updates)} WHERE id = ?"
                cursor.execute(query, params)
                conn.commit()
                return True
        except Exception as e:
            print(f"Güncelleme hatası: {e}")
            return False
    
    def get_summary(self, kisi=None, yil=None):
        """Özet istatistikleri getir - her iki tablodan"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # AXA tablosu
            query_axa = """
            SELECT 
                COUNT(*) as kayit_sayisi,
                COALESCE(SUM(brut_prim), 0) as toplam_brut,
                COALESCE(SUM(net_prim), 0) as toplam_net,
                COALESCE(SUM(toplam_komisyon), 0) as toplam_komisyon,
                COALESCE(SUM(odenen_komisyon), 0) as toplam_odenen
            FROM komisyonlar_axa WHERE 1=1
            """
            params = []
            
            if kisi:
                query_axa += " AND kisi = ?"
                params.append(kisi)
            if yil:
                query_axa += " AND yil = ?"
                params.append(yil)
            
            cursor.execute(query_axa, params)
            result_axa = cursor.fetchone()
            
            # Diğer şirketler tablosu
            query_other = """
            SELECT 
                COUNT(*) as kayit_sayisi,
                COALESCE(SUM(brut_prim), 0) as toplam_brut,
                COALESCE(SUM(net_prim), 0) as toplam_net,
                COALESCE(SUM(toplam_komisyon), 0) as toplam_komisyon,
                COALESCE(SUM(odenen_komisyon), 0) as toplam_odenen
            FROM komisyonlar_other WHERE 1=1
            """
            params_other = []
            
            if kisi:
                query_other += " AND kisi = ?"
                params_other.append(kisi)
            if yil:
                query_other += " AND yil = ?"
                params_other.append(yil)
            
            cursor.execute(query_other, params_other)
            result_other = cursor.fetchone()
            
            conn.close()
            
            # Toplamları birleştir
            return {
                'kayit_sayisi': result_axa[0] + result_other[0],
                'toplam_brut': result_axa[1] + result_other[1],
                'toplam_net': result_axa[2] + result_other[2],
                'toplam_komisyon': result_axa[3] + result_other[3],
                'toplam_odenen': result_axa[4] + result_other[4]
            }
        except Exception as e:
            print(f"Özet hatası: {e}")
            return {'kayit_sayisi': 0, 'toplam_brut': 0, 'toplam_net': 0, 'toplam_komisyon': 0, 'toplam_odenen': 0}

# ==============================================================================
# 🖥️ ANA GUI SINIFI
# ==============================================================================

class MuhasebeGUI:
    """Ana GUI Sınıfı"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("🏢 AXA Muhasebe ve Komisyon Takip Sistemi")
        self.root.geometry("1500x850")
        
        # Veritabanı
        self.db = MuhasebeDB()
        
        # Mevcut seçimler
        self.current_kisi = "YAŞAR"
        self.current_yil = datetime.now().year
        
        # UI oluştur
        self.create_ui()
        
        # Başlangıç
        self.log("✅ Uygulama başlatıldı", "success")
        self.load_data()
        self.update_summary()
    
    def create_ui(self):
        """Kullanıcı arayüzünü oluştur"""
        self.create_menu()
        self.create_header()
        self.create_tabs()
        self.create_content()
        self.create_status_bar()
    
    def create_menu(self):
        """Menü çubuğu"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # Dosya
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="📁 Dosya", menu=file_menu)
        file_menu.add_command(label="PDF Klasörü İşle", command=self.process_pdf_folder)
        file_menu.add_command(label="Excel'den İçe Aktar", command=self.import_excel_data)
        file_menu.add_command(label="Excel'e Aktar", command=self.export_to_excel)
        file_menu.add_separator()
        file_menu.add_command(label="Çıkış", command=self.root.quit)
        
        # Düzenle
        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="✏️ Düzenle", menu=edit_menu)
        edit_menu.add_command(label="Yeni Kayıt Ekle", command=self.add_manual_entry)
        edit_menu.add_command(label="Seçili Kaydı Sil", command=self.delete_selected)
        
        # Görünüm
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="👁️ Görünüm", menu=view_menu)
        view_menu.add_command(label="Yenile", command=self.load_data)
        view_menu.add_command(label="Logları Temizle", command=self.clear_logs)
        
        # Yardım
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="❓ Yardım", menu=help_menu)
        help_menu.add_command(label="Hakkında", command=self.show_about)
    
    def create_header(self):
        """Başlık bölümü"""
        header = tk.Frame(self.root, bg=COLORS['primary'], height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        title = tk.Label(
            header,
            text="🏢 AXA MUHASEBE VE KOMİSYON TAKİP SİSTEMİ",
            font=('Arial', 14, 'bold'),
            bg=COLORS['primary'],
            fg='white'
        )
        title.pack(pady=5)
        
        subtitle = tk.Label(
            header,
            text="Otomatik PDF Okuma ve Komisyon Hesaplama Sistemi",
            font=('Arial', 9),
            bg=COLORS['primary'],
            fg='white'
        )
        subtitle.pack()
    
    def create_tabs(self):
        """Kişi tab'ları"""
        tab_frame = tk.Frame(self.root, bg='#e2e8f0')
        tab_frame.pack(fill=tk.X)
        
        self.tab_buttons = {}
        for kisi in KISILER:
            btn = tk.Button(
                tab_frame,
                text=kisi,
                font=('Arial', 9, 'bold'),
                bg='white' if kisi == self.current_kisi else '#e2e8f0',
                fg=COLORS['primary'] if kisi == self.current_kisi else '#64748b',
                relief=tk.FLAT,
                padx=20,
                pady=3,
                cursor='hand2',
                command=lambda k=kisi: self.switch_tab(k)
            )
            btn.pack(side=tk.LEFT, padx=1, pady=2)
            self.tab_buttons[kisi] = btn
    
    def switch_tab(self, kisi):
        """Tab değiştir"""
        self.current_kisi = kisi
        
        # Buton görünümlerini güncelle
        for k, btn in self.tab_buttons.items():
            if k == kisi:
                btn.config(bg='white', fg=COLORS['primary'])
            else:
                btn.config(bg='#e2e8f0', fg='#64748b')
        
        self.load_data()
        self.update_summary()
        self.log(f"📂 {kisi} sekmesine geçildi", "info")
    
    def create_content(self):
        """Ana içerik alanı"""
        main_frame = tk.Frame(self.root, bg=COLORS['bg_light'])
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Sol panel (darı tutuldu)
        left_panel = tk.Frame(main_frame, bg='white', width=150)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        left_panel.pack_propagate(False)
        
        self.create_filter_panel(left_panel)
        self.create_buttons(left_panel)
        self.create_summary_panel(left_panel)
        self.create_log_panel(left_panel)  # Log paneli sol panele taşındı
        
        # Sağ panel
        right_panel = tk.Frame(main_frame, bg=COLORS['bg_light'])
        right_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.create_table(right_panel)
    
    def create_filter_panel(self, parent):
        """Filtre paneli"""
        filter_frame = tk.LabelFrame(
            parent,
            text="🔍 FİLTRELER",
            font=('Arial', 9, 'bold'),
            bg='white',
            fg=COLORS['primary']
        )
        filter_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Yıl
        tk.Label(filter_frame, text="Yıl:", font=('Arial', 8), bg='white').pack(anchor='w', padx=5, pady=(5, 0))
        years = [str(y) for y in range(2020, datetime.now().year + 2)]
        self.filter_yil = ttk.Combobox(filter_frame, values=['Tümü'] + years, font=('Arial', 8), state='readonly')
        self.filter_yil.set('Tümü')
        self.filter_yil.pack(fill=tk.X, padx=5, pady=(1, 3))
        self.filter_yil.bind('<<ComboboxSelected>>', lambda e: self.on_filter_change())
        
        # Sigortalı
        tk.Label(filter_frame, text="Sigortalı:", font=('Arial', 8), bg='white').pack(anchor='w', padx=5, pady=(2, 0))
        self.filter_sigortali = tk.Entry(filter_frame, font=('Arial', 8))
        self.filter_sigortali.pack(fill=tk.X, padx=5, pady=(1, 5))
        
        # Ara butonu
        search_btn = tk.Button(
            filter_frame,
            text="🔍 ARA",
            command=self.apply_filters,
            font=('Arial', 8, 'bold'),
            bg=COLORS['primary'],
            fg='white',
            relief=tk.FLAT,
            cursor='hand2',
            pady=4
        )
        search_btn.pack(fill=tk.X, padx=5, pady=(2, 5))
    
    def create_buttons(self, parent):
        """Butonlar"""
        btn_frame = tk.LabelFrame(
            parent,
            text="⚡ İŞLEMLER",
            font=('Arial', 9, 'bold'),
            bg='white'
        )
        btn_frame.pack(fill=tk.X, padx=5, pady=5)
        
        buttons = [
            ("📂 PDF İşle", self.process_pdf_folder, '#2563eb'),
            ("📊 Exceli Aç", self.import_excel_data, '#8b5cf6'),
            ("💾 Excel'e Aktar", self.export_to_excel, '#ea580c'),
            ("🗑️ Hepsini Sil", self.delete_all_records, COLORS['danger'])
        ]
        
        for text, command, color in buttons:
            btn = tk.Button(
                btn_frame,
                text=text,
                command=command,
                font=('Arial', 8, 'bold'),
                bg=color,
                fg='white',
                relief=tk.FLAT,
                cursor='hand2',
                padx=5,
                pady=4
            )
            btn.pack(fill=tk.X, padx=5, pady=2)
    
    def create_summary_panel(self, parent):
        """Özet paneli"""
        summary_frame = tk.LabelFrame(
            parent,
            text="📈 ÖZET",
            font=('Arial', 9, 'bold'),
            bg='white'
        )
        summary_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.summary_text = tk.Text(
            summary_frame,
            font=('Consolas', 8),
            bg='#f8fafc',
            fg='#1e293b',
            relief=tk.FLAT,
            wrap=tk.WORD
        )
        self.summary_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
    
    def create_table(self, parent):
        """İkili kayıt tablosu - Sol: AXA, Sağ: Diğer Şirketler"""
        self.table_frame = tk.Frame(parent, bg='white')
        self.table_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # İki panel için PanedWindow
        self.paned = tk.PanedWindow(self.table_frame, orient=tk.HORIZONTAL, bg='#e2e8f0', sashwidth=4)
        self.paned.pack(fill=tk.BOTH, expand=True)
        
        # ============ SOL PANEL: AXA POLİÇELERİ ============
        left_frame = tk.Frame(self.paned, bg='#fff3cd')
        
        # Sol panel başlık ve + butonu
        left_header = tk.Frame(left_frame, bg='#fff3cd')
        left_header.pack(fill=tk.X)
        
        tk.Label(left_header, text="📋 AXA POLİÇELERİ", font=('Arial', 10, 'bold'),
                 bg='#fff3cd', fg='#856404').pack(side=tk.LEFT, padx=5, pady=2)
        
        self.btn_add_axa = tk.Button(
            left_header, text="+", font=('Arial', 12, 'bold'),
            bg='#16a34a', fg='white', width=2, cursor='hand2',
            command=lambda: self.add_manual_entry_to_table('axa')
        )
        self.btn_add_axa.pack(side=tk.RIGHT, padx=5, pady=2)
        
        # Sol tablo scrollbar
        left_scroll_y = ttk.Scrollbar(left_frame, orient=tk.VERTICAL)
        left_scroll_x = ttk.Scrollbar(left_frame, orient=tk.HORIZONTAL)
        
        # Sol sütunlar (şirket yok, PLAKA yok)
        columns_axa = ('ID', 'SİGORTALI', 'TARİH', 'POLİÇE_NO', 'TÜR', 
                       'ÖDEME', 'BRÜT_PRİM', 'NET_PRİM', 'KOMİSYON', 'ÖDENEN')
        
        self.tree_axa = ttk.Treeview(
            left_frame,
            columns=columns_axa,
            show='headings',
            yscrollcommand=left_scroll_y.set,
            xscrollcommand=left_scroll_x.set,
            selectmode='extended'
        )
        
        headers_axa = {
            'ID': 'ID', 'SİGORTALI': 'SİGORTALI', 'TARİH': 'TARİH',
            'POLİÇE_NO': 'POLİÇE NO', 'TÜR': 'TÜR',
            'ÖDEME': 'ÖDEME', 'BRÜT_PRİM': 'BRÜT PRİM', 'NET_PRİM': 'NET PRİM',
            'KOMİSYON': 'KOMİSYON', 'ÖDENEN': 'ÖDENEN'
        }
        
        widths_axa = {
            'ID': 30, 'SİGORTALI': 130, 'TARİH': 70, 'POLİÇE_NO': 80,
            'TÜR': 55, 'ÖDEME': 25, 'BRÜT_PRİM': 75,
            'NET_PRİM': 70, 'KOMİSYON': 65, 'ÖDENEN': 65
        }
        
        for col in columns_axa:
            self.tree_axa.heading(col, text=headers_axa[col])
            self.tree_axa.column(col, width=widths_axa[col], anchor='center' if col != 'SİGORTALI' else 'w')
        
        left_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        left_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree_axa.pack(fill=tk.BOTH, expand=True)
        
        left_scroll_y.config(command=self.tree_axa.yview)
        left_scroll_x.config(command=self.tree_axa.xview)
        
        # ============ SAĞ PANEL: DİĞER ŞİRKETLER ============
        right_frame = tk.Frame(self.paned, bg='#d1ecf1')
        
        # Sağ panel başlık ve + butonu
        right_header = tk.Frame(right_frame, bg='#d1ecf1')
        right_header.pack(fill=tk.X)
        
        tk.Label(right_header, text="🏢 DİĞER ŞİRKETLER", font=('Arial', 10, 'bold'),
                 bg='#d1ecf1', fg='#0c5460').pack(side=tk.LEFT, padx=5, pady=2)
        
        self.btn_add_other = tk.Button(
            right_header, text="+", font=('Arial', 12, 'bold'),
            bg='#16a34a', fg='white', width=2, cursor='hand2',
            command=lambda: self.add_manual_entry_to_table('other')
        )
        self.btn_add_other.pack(side=tk.RIGHT, padx=5, pady=2)
        
        # Sağ tablo scrollbar
        right_scroll_y = ttk.Scrollbar(right_frame, orient=tk.VERTICAL)
        right_scroll_x = ttk.Scrollbar(right_frame, orient=tk.HORIZONTAL)
        
        # Sağ sütunlar (şirket VAR, PLAKA yok)
        columns_other = ('ID', 'ŞİRKET', 'SİGORTALI', 'TARİH', 'POLİÇE_NO', 'TÜR', 
                         'ÖDEME', 'BRÜT_PRİM', 'NET_PRİM', 'KOMİSYON', 'ÖDENEN')
        
        self.tree_other = ttk.Treeview(
            right_frame,
            columns=columns_other,
            show='headings',
            yscrollcommand=right_scroll_y.set,
            xscrollcommand=right_scroll_x.set,
            selectmode='extended'
        )
        
        headers_other = {
            'ID': 'ID', 'ŞİRKET': 'ŞİRKET', 'SİGORTALI': 'SİGORTALI', 'TARİH': 'TARİH',
            'POLİÇE_NO': 'POLİÇE NO', 'TÜR': 'TÜR',
            'ÖDEME': 'ÖDEME', 'BRÜT_PRİM': 'BRÜT PRİM', 'NET_PRİM': 'NET PRİM',
            'KOMİSYON': 'KOMİSYON', 'ÖDENEN': 'ÖDENEN'
        }
        
        widths_other = {
            'ID': 30, 'ŞİRKET': 55, 'SİGORTALI': 120, 'TARİH': 70, 'POLİÇE_NO': 80,
            'TÜR': 55, 'ÖDEME': 25, 'BRÜT_PRİM': 75,
            'NET_PRİM': 70, 'KOMİSYON': 65, 'ÖDENEN': 65
        }
        
        for col in columns_other:
            self.tree_other.heading(col, text=headers_other[col])
            self.tree_other.column(col, width=widths_other[col], anchor='center' if col not in ['SİGORTALI', 'ŞİRKET'] else 'w')
        
        right_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        right_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree_other.pack(fill=tk.BOTH, expand=True)
        
        right_scroll_y.config(command=self.tree_other.yview)
        right_scroll_x.config(command=self.tree_other.xview)
        
        # Panelleri eşit boyutta ekle (stretch=always ile %50-%50)
        self.paned.add(left_frame, stretch='always')
        self.paned.add(right_frame, stretch='always')
        
        # Ana tree referansı (uyumluluk için - AXA tablosu)
        self.tree = self.tree_axa
        
        # Çift tık düzenleme - her iki tablo için
        self.tree_axa.bind('<Double-Button-1>', self.on_double_click)
        self.tree_other.bind('<Double-Button-1>', self.on_double_click_other)
        
        # Sağ tık menü
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="🗑️ Sil", command=self.delete_selected)
        self.tree_axa.bind('<Button-3>', self.show_context_menu)
        self.tree_other.bind('<Button-3>', self.show_context_menu_other)
        
        # Inline editing için değişkenler
        self.editing_new_row = False
        self.new_row_item = None
        self.current_edit_entry = None
        self.active_tree = self.tree_axa  # Aktif tablo
    
    def create_log_panel(self, parent):
        """Log paneli (küçük, özet altında)"""
        log_frame = tk.LabelFrame(
            parent,
            text="📋 LOG",
            font=('Arial', 8, 'bold'),
            bg='white'
        )
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.log_text = tk.Text(
            log_frame,
            height=4,
            font=('Consolas', 7),
            bg='#fafafa',
            fg='#374151',
            wrap=tk.WORD
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)
        
        self.log_text.tag_config('success', foreground='#16a34a')
        self.log_text.tag_config('error', foreground='#dc2626')
        self.log_text.tag_config('warning', foreground='#ea580c')
        self.log_text.tag_config('info', foreground='#2563eb')
    
    def create_status_bar(self):
        """Durum çubuğu"""
        self.status_bar = tk.Label(
            self.root,
            text="Hazır",
            font=('Arial', 9),
            bg='#e2e8f0',
            fg='#475569',
            anchor='w',
            padx=10
        )
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
    
    # =========================================================================
    # VERİ İŞLEMLERİ
    # =========================================================================
    
    def on_filter_change(self):
        """Filtre değiştiğinde"""
        yil = self.filter_yil.get()
        self.current_yil = int(yil) if yil != 'Tümü' else None
        
        self.load_data()
        self.update_summary()
    
    def apply_filters(self):
        """Filtreleri uygula"""
        self.load_data()
        self.update_summary()
    
    def load_data(self):
        """Verileri yükle - AXA ve diğer şirketleri ayrı tablolara"""
        # Her iki tabloyu temizle
        for item in self.tree_axa.get_children():
            self.tree_axa.delete(item)
        for item in self.tree_other.get_children():
            self.tree_other.delete(item)
        
        yil = int(self.filter_yil.get()) if self.filter_yil.get() != 'Tümü' else None
        
        df = self.db.get_records(kisi=self.current_kisi, yil=yil)
        
        # Sigortalı filtresi
        sigortali_filter = self.filter_sigortali.get().strip().upper()
        if sigortali_filter and not df.empty:
            df = df[df['sigortali'].str.upper().str.contains(sigortali_filter, na=False)]
        
        if df.empty:
            self.log("ℹ️ Gösterilecek kayıt yok", "info")
            return
        
        axa_count = 0
        other_count = 0
        
        for _, row in df.iterrows():
            # Ödeme türü - işaret ekle
            odeme = str(row['odeme_turu']).upper().strip()
            if 'K.KART' in odeme or 'KART' in odeme:
                odeme_display = '●'
            else:
                odeme_display = '■'
            
            # Şirket bilgisi - varsayılan AXA
            sirket = str(row.get('sirket', 'AXA')).upper().strip()
            if not sirket or sirket == 'NONE' or sirket == 'NAN':
                sirket = 'AXA'
            
            if sirket == 'AXA':
                # AXA tablosuna ekle (şirket ve PLAKA sütunu yok)
                values_axa = (
                    row['id'],
                    row['sigortali'],
                    row['tarih'],
                    row['police_no'],
                    row['tur'],
                    odeme_display,
                    f"{row['brut_prim']:,.2f}",
                    f"{row['net_prim']:,.2f}",
                    f"{row['toplam_komisyon']:,.2f}",
                    f"{row['odenen_komisyon']:,.2f}"
                )
                self.tree_axa.insert('', tk.END, values=values_axa)
                axa_count += 1
            else:
                # Diğer şirketler tablosuna ekle (şirket VAR, PLAKA yok)
                values_other = (
                    row['id'],
                    sirket,
                    row['sigortali'],
                    row['tarih'],
                    row['police_no'],
                    row['tur'],
                    odeme_display,
                    f"{row['brut_prim']:,.2f}",
                    f"{row['net_prim']:,.2f}",
                    f"{row['toplam_komisyon']:,.2f}",
                    f"{row['odenen_komisyon']:,.2f}"
                )
                self.tree_other.insert('', tk.END, values=values_other)
                other_count += 1
        
        self.log(f"✅ AXA: {axa_count}, Diğer: {other_count} kayıt yüklendi", "success")
    
    def update_summary(self):
        """Özeti güncelle"""
        yil = int(self.filter_yil.get()) if self.filter_yil.get() != 'Tümü' else None
        
        summary = self.db.get_summary(kisi=self.current_kisi, yil=yil)
        
        text = f"📊 {self.current_kisi}\n"
        text += "=" * 25 + "\n\n"
        text += f"📄 Kayıt: {summary['kayit_sayisi']}\n\n"
        text += f"💰 Brüt: {summary['toplam_brut']:,.2f} TL\n"
        text += f"💵 Net: {summary['toplam_net']:,.2f} TL\n\n"
        text += f"📈 Komisyon:\n"
        text += f"   {summary['toplam_komisyon']:,.2f} TL\n\n"
        text += f"💳 Ödenen:\n"
        text += f"   {summary['toplam_odenen']:,.2f} TL\n"
        
        self.summary_text.delete('1.0', tk.END)
        self.summary_text.insert('1.0', text)
    
    # =========================================================================
    # PDF İŞLEME
    # =========================================================================
    
    def process_pdf_folder(self):
        """PDF klasörü işle"""
        if not PARSER_AVAILABLE:
            messagebox.showerror("Hata", "axa_parser.py modülü bulunamadı!")
            return
        
        folder = filedialog.askdirectory(title="PDF Klasörünü Seçin")
        if not folder:
            return
        
        pdf_files = [f for f in os.listdir(folder) if f.lower().endswith('.pdf')]
        
        if not pdf_files:
            messagebox.showwarning("Uyarı", "Klasörde PDF dosyası bulunamadı!")
            return
        
        self.log(f"📂 {len(pdf_files)} PDF dosyası bulundu", "info")
        
        # Progress
        progress_win = tk.Frame(self.root, bg='white', relief=tk.RIDGE, borderwidth=2)
        progress_win.place(relx=0.5, rely=0.5, anchor='center', width=400, height=120)
        
        tk.Label(progress_win, text="📄 PDF İşleniyor...", font=('Arial', 11, 'bold'), bg='white').pack(pady=10)
        progress_label = tk.Label(progress_win, text="", font=('Arial', 9), bg='white')
        progress_label.pack(pady=5)
        progress_bar = ttk.Progressbar(progress_win, length=350, mode='determinate')
        progress_bar.pack(pady=10)
        
        basarili = 0
        basarisiz = 0
        duplicate = 0
        
        for idx, pdf_file in enumerate(pdf_files):
            pdf_path = os.path.join(folder, pdf_file)
            progress_label.config(text=pdf_file[:40])
            progress_bar['value'] = ((idx + 1) / len(pdf_files)) * 100
            progress_win.update()
            
            try:
                data = process_pdf(pdf_path, pdf_file)
                
                if data:
                    # Tutar parse
                    brut_prim = parse_turkish_amount(data.get('TUTAR', '0'))
                    
                    # Komisyon hesapla
                    komisyon = hesapla_komisyon(self.current_kisi, data.get('TÜR', ''), brut_prim)
                    
                    # Tarihten yıl çıkar
                    tarih = data.get('TARİH', '')
                    yil = datetime.now().year
                    if tarih and '/' in tarih:
                        try:
                            yil = int(tarih.split('/')[2])
                        except:
                            pass
                    
                    record = {
                        'kisi': self.current_kisi,
                        'ay': '-',
                        'yil': yil,
                        'sigortali': data.get('SİGORTALI', '-'),
                        'police_no': data.get('POLİÇE NO', '-'),
                        'tarih': tarih,
                        'plaka': data.get('PLAKA', '-'),
                        'tur': data.get('TÜR', '-'),
                        'odeme_turu': 'AÇIK',
                        'brut_prim': brut_prim,
                        'net_prim': komisyon['net_prim'],
                        'komisyon_orani': komisyon['komisyon_orani'],
                        'toplam_komisyon': komisyon['toplam_komisyon'],
                        'odeme_orani': komisyon['odeme_orani'],
                        'odenen_komisyon': komisyon['odenen_komisyon'],
                        'sirket': data.get('ŞİRKET', 'AXA')
                    }
                    
                    if self.db.insert_record(record):
                        basarili += 1
                        sirket = data.get('ŞİRKET', 'AXA')
                        self.log(f"✅ {pdf_file} - {sirket} - {data.get('TÜR', '?')}", "success")
                    else:
                        duplicate += 1
                        self.log(f"⚠️ Duplicate: {pdf_file}", "warning")
                else:
                    basarisiz += 1
                    self.log(f"⚠️ İşlenemedi: {pdf_file}", "warning")
            
            except Exception as e:
                basarisiz += 1
                self.log(f"❌ Hata ({pdf_file}): {str(e)}", "error")
        
        progress_win.destroy()
        
        msg = f"✅ Başarılı: {basarili}\n"
        if duplicate > 0:
            msg += f"⚠️ Duplicate: {duplicate}\n"
        if basarisiz > 0:
            msg += f"❌ Hatalı: {basarisiz}"
        
        messagebox.showinfo("Tamamlandı", msg)
        self.load_data()
        self.update_summary()
    
    # =========================================================================
    # MANUEL KAYIT
    # =========================================================================
    
    def add_manual_entry(self):
        """Eski fonksiyon - geriye dönük uyumluluk için"""
        self.add_manual_entry_to_table('axa')
    
    def add_manual_entry_to_table(self, table_type='axa'):
        """Manuel kayıt ekle - tablonun en üstüne düzenlenebilir boş satır ekle"""
        # Zaten düzenleme modundaysa uyar
        if self.editing_new_row:
            messagebox.showwarning("Uyarı", "Lütfen önce mevcut satırı tamamlayın veya silin!")
            return
        
        # Hangi tabloyu kullanacağız?
        if table_type == 'other':
            self.active_tree = self.tree_other
            # Sağ tablo için sütunlar (ŞİRKET var)
            values = ('🆕', 'AXA', '(Sigortalı)', '(Tarih)', '(Poliçe No)', '(Tür)', 'AÇIK', '(Brüt Prim)', '-', '-', '-')
        else:
            self.active_tree = self.tree_axa
            # Sol tablo için sütunlar (ŞİRKET yok)
            values = ('🆕', '(Sigortalı)', '(Tarih)', '(Poliçe No)', '(Tür)', 'AÇIK', '(Brüt Prim)', '-', '-', '-')
        
        # Tablonun en üstüne boş satır ekle
        self.new_row_item = self.active_tree.insert('', 0, values=values, tags=('new_row',))
        
        # Yeni satır için özel stil
        self.active_tree.tag_configure('new_row', background='#e8f4fd')
        
        self.editing_new_row = True
        self.current_table_type = table_type
        self.new_row_data = {
            'sirket': 'AXA' if table_type == 'other' else 'AXA',
            'sigortali': '',
            'tarih': '',
            'police_no': '',
            'plaka': '-',
            'tur': '',
            'odeme': 'AÇIK',
            'brut_prim': '',
            'notlar': ''
        }
        
        # Yeni satırı seç ve görünür yap
        self.active_tree.selection_set(self.new_row_item)
        self.active_tree.see(self.new_row_item)
        
        self.log("➕ Yeni satır eklendi. Hücrelere çift tıklayarak doldurun ve Enter ile kaydedin.", "info")
        
        # Çift tıklama event'ini güncelle - sadece aktif tablo için
        self.active_tree.bind('<Double-Button-1>', self.on_double_click_new_row)
        # Diğer tablonun event'i aynı kalsın
        if self.active_tree == self.tree_axa:
            self.tree_other.bind('<Double-Button-1>', self.on_double_click_other)
        else:
            self.tree_axa.bind('<Double-Button-1>', self.on_double_click)
    
    def on_double_click_new_row(self, event):
        """Yeni satır için çift tıklama - inline editing"""
        region = self.active_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        column = self.active_tree.identify_column(event.x)
        item = self.active_tree.identify_row(event.y)
        
        if not item:
            return
        
        # Eğer yeni satır değilse normal çift tık davranışı
        if item != self.new_row_item:
            if self.active_tree == self.tree_other:
                self.on_double_click_other(event)
            else:
                self.on_double_click(event)
            return
        
        col_index = int(column.replace('#', '')) - 1
        # Sütun listesi tabloya göre farklı
        if self.current_table_type == 'other':
            columns_list = ['ID', 'ŞİRKET', 'SİGORTALI', 'TARİH', 'POLİÇE_NO', 'TÜR', 
                            'ÖDEME', 'BRÜT_PRİM', 'NET_PRİM', 'KOMİSYON', 'ÖDENEN']
        else:
            columns_list = ['ID', 'SİGORTALI', 'TARİH', 'POLİÇE_NO', 'TÜR', 
                            'ÖDEME', 'BRÜT_PRİM', 'NET_PRİM', 'KOMİSYON', 'ÖDENEN']
        
        if col_index >= len(columns_list):
            return
        
        col_name = columns_list[col_index]
        
        # Düzenlenebilir sütunlar
        editable_cols = {
            'ŞİRKET': 'sirket',
            'SİGORTALI': 'sigortali',
            'TARİH': 'tarih', 
            'POLİÇE_NO': 'police_no',
            'TÜR': 'tur',
            'ÖDEME': 'odeme',
            'BRÜT_PRİM': 'brut_prim'
        }
        
        if col_name not in editable_cols:
            return
        
        # Önceki entry'yi kapat
        if self.current_edit_entry:
            self.current_edit_entry.destroy()
        
        # Hücre koordinatlarını al
        bbox = self.active_tree.bbox(item, column)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        field_key = editable_cols[col_name]
        
        # ŞİRKET, TÜR ve ÖDEME için Combobox, diğerleri için Entry
        if col_name == 'ŞİRKET':
            sirketler = ['AXA', 'ALLIANZ', 'ANADOLU', 'GÜNEŞ', 'MAPFRE', 'QUICK', 'HDI', 'DOĞA', 'SOMPİ', 'RAY']
            self.current_edit_entry = ttk.Combobox(self.active_tree, values=sirketler, 
                                                    font=('Arial', 9), state='normal', width=10)
            self.current_edit_entry.set(self.new_row_data.get('sirket', 'AXA') or 'AXA')
        elif col_name == 'TÜR':
            self.current_edit_entry = ttk.Combobox(self.active_tree, values=TURLER, 
                                                    font=('Arial', 9), state='readonly', width=8)
            self.current_edit_entry.set(self.new_row_data.get('tur', TURLER[0]) or TURLER[0])
        elif col_name == 'ÖDEME':
            self.current_edit_entry = ttk.Combobox(self.active_tree, values=ODEME_TURLERI, 
                                                    font=('Arial', 9), state='readonly', width=6)
            self.current_edit_entry.set(self.new_row_data.get('odeme', 'AÇIK'))
        else:
            self.current_edit_entry = tk.Entry(self.active_tree, font=('Arial', 9))
            current_val = self.new_row_data.get(field_key, '')
            if current_val and not current_val.startswith('('):
                self.current_edit_entry.insert(0, current_val)
        
        self.current_edit_entry.place(x=x, y=y, width=width, height=height)
        self.current_edit_entry.focus_set()
        
        # Enter ile kaydet, FocusOut ile de kaydet
        self.current_edit_entry.bind('<Return>', lambda e: self.save_cell_edit(field_key, col_index))
        self.current_edit_entry.bind('<FocusOut>', lambda e: self.save_cell_edit(field_key, col_index))
        
        # Combobox için seçim event'i
        if isinstance(self.current_edit_entry, ttk.Combobox):
            self.current_edit_entry.bind('<<ComboboxSelected>>', lambda e: self.save_cell_edit(field_key, col_index))
    
    def save_cell_edit(self, field_key, col_index, move_next=False):
        """Hücre düzenlemesini kaydet"""
        if not self.current_edit_entry:
            return
        
        value = self.current_edit_entry.get().strip()
        self.new_row_data[field_key] = value
        
        # Treeview'daki değeri güncelle
        current_values = list(self.active_tree.item(self.new_row_item, 'values'))
        
        # Değeri güncelle - tablo tipine göre farklı indeksler
        if self.current_table_type == 'other':
            # Sağ tablo: ID, ŞİRKET, SİGORTALI, TARİH, POLİÇE_NO, TÜR, ÖDEME, BRÜT_PRİM, NET, KOM, ÖDENEN
            if field_key == 'sirket':
                current_values[1] = value if value else 'AXA'
            elif field_key == 'sigortali':
                current_values[2] = value if value else '(Sigortalı)'
            elif field_key == 'tarih':
                current_values[3] = value if value else '(Tarih)'
            elif field_key == 'police_no':
                current_values[4] = value if value else '(Poliçe No)'
            elif field_key == 'tur':
                current_values[5] = value if value else '(Tür)'
            elif field_key == 'odeme':
                current_values[6] = value if value else 'AÇIK'
            elif field_key == 'brut_prim':
                current_values[7] = value if value else '(Brüt Prim)'
        else:
            # Sol tablo: ID, SİGORTALI, TARİH, POLİÇE_NO, TÜR, ÖDEME, BRÜT_PRİM, NET, KOM, ÖDENEN
            if field_key == 'sigortali':
                current_values[1] = value if value else '(Sigortalı)'
            elif field_key == 'tarih':
                current_values[2] = value if value else '(Tarih)'
            elif field_key == 'police_no':
                current_values[3] = value if value else '(Poliçe No)'
            elif field_key == 'tur':
                current_values[4] = value if value else '(Tür)'
            elif field_key == 'odeme':
                current_values[5] = value if value else 'AÇIK'
            elif field_key == 'brut_prim':
                current_values[6] = value if value else '(Brüt Prim)'
        
        self.active_tree.item(self.new_row_item, values=current_values)
        
        # Entry'yi kapat
        if self.current_edit_entry:
            self.current_edit_entry.destroy()
            self.current_edit_entry = None
    
    def cancel_cell_edit(self):
        """Hücre düzenlemesini iptal et"""
        if self.current_edit_entry:
            self.current_edit_entry.destroy()
            self.current_edit_entry = None
    
    def cancel_edit_and_remove(self):
        """Düzenlemeyi iptal et ve yeni satırı sil"""
        if self.current_edit_entry:
            self.current_edit_entry.destroy()
            self.current_edit_entry = None
        
        if self.new_row_item and self.editing_new_row:
            self.active_tree.delete(self.new_row_item)
            self.editing_new_row = False
            self.new_row_item = None
            # Her iki tablonun event'ini de geri yükle
            self.tree_axa.bind('<Double-Button-1>', self.on_double_click)
            self.tree_other.bind('<Double-Button-1>', self.on_double_click_other)
            self.log("❌ Yeni kayıt ekleme iptal edildi", "warning")
    
    def check_and_save_new_row(self):
        """Girilen değerlerle kayıt ekle (zorunlu alan kontrolü yok)"""
        # Brüt prim kontrolü
        brut_prim_str = self.new_row_data.get('brut_prim', '0')
        if brut_prim_str and not brut_prim_str.startswith('('):
            try:
                brut_prim = float(brut_prim_str.replace('.', '').replace(',', '.'))
            except:
                brut_prim = 0
        else:
            brut_prim = 0
        
        tarih = self.new_row_data.get('tarih', '')
        if tarih and tarih.startswith('('):
            tarih = datetime.now().strftime('%d/%m/%Y')
        
        yil = datetime.now().year
        if tarih and '/' in tarih:
            try:
                yil = int(tarih.split('/')[2])
            except:
                pass
        
        tur = self.new_row_data.get('tur', '-')
        if tur.startswith('('):
            tur = '-'
        
        # Komisyon hesapla (sadece geçerli tür varsa)
        if tur and tur != '-' and brut_prim > 0:
            komisyon = hesapla_komisyon(self.current_kisi, tur, brut_prim)
        else:
            komisyon = {
                'net_prim': brut_prim,
                'komisyon_orani': 0,
                'toplam_komisyon': 0,
                'odeme_orani': 0,
                'odenen_komisyon': 0
            }
        
        # Poliçe no kontrolü
        police_no = self.new_row_data.get('police_no', '')
        if not police_no or police_no.startswith('('):
            police_no = f"MNL-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        sigortali = self.new_row_data.get('sigortali', '-')
        if sigortali.startswith('('):
            sigortali = '-'
        
        record = {
            'kisi': self.current_kisi,
            'ay': '-',
            'yil': yil,
            'sigortali': sigortali,
            'police_no': police_no,
            'tarih': tarih,
            'plaka': '-',
            'tur': tur,
            'odeme_turu': self.new_row_data.get('odeme', 'AÇIK'),
            'brut_prim': brut_prim,
            'net_prim': komisyon['net_prim'],
            'komisyon_orani': komisyon['komisyon_orani'],
            'toplam_komisyon': komisyon['toplam_komisyon'],
            'odeme_orani': komisyon['odeme_orani'],
            'odenen_komisyon': komisyon['odenen_komisyon'],
            'notlar': self.new_row_data.get('notlar', ''),
            'sirket': self.new_row_data.get('sirket', 'AXA')
        }
        
        if self.db.insert_record(record):
            self.log("✅ Yeni kayıt başarıyla eklendi!", "success")
            self.editing_new_row = False
            self.new_row_item = None
            # Her iki tablonun event'ini de geri yükle
            self.tree_axa.bind('<Double-Button-1>', self.on_double_click)
            self.tree_other.bind('<Double-Button-1>', self.on_double_click_other)
            self.load_data()
            self.update_summary()
        else:
            self.log("❌ Kayıt eklenemedi! (Poliçe no zaten mevcut olabilir)", "error")
    
    # =========================================================================
    # DİĞER İŞLEMLER
    # =========================================================================
    
    def on_double_click(self, event, table_type='axa'):
        """Çift tıklama - hücreyi düzenle"""
        # Hangi tablo kullanılacak
        if table_type == 'other':
            tree = self.tree_other
        else:
            tree = self.tree_axa
        
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        column = tree.identify_column(event.x)
        item = tree.identify_row(event.y)
        
        if not item:
            return
        
        col_index = int(column.replace('#', '')) - 1
        # Sütun listesi güncellendi (PLAKA kaldırıldı)
        columns = ('ID', 'SİGORTALI', 'TARİH', 'POLİÇE_NO', 'TÜR', 
                   'ÖDEME', 'BRÜT_PRİM', 'NET_PRİM', 'KOMİSYON', 'ÖDENEN')
        
        if col_index >= len(columns):
            return
        
        col_name = columns[col_index]
        values = tree.item(item, 'values')
        record_id = values[0]
        
        # Düzenlenebilir sütunlar ve veritabanı alanları
        editable_cols = {
            'SİGORTALI': ('sigortali', 'entry'),
            'TARİH': ('tarih', 'entry'),
            'POLİÇE_NO': ('police_no', 'entry'),
            'TÜR': ('tur', 'combo_tur'),
            'ÖDEME': ('odeme_turu', 'combo_odeme'),
            'BRÜT_PRİM': ('brut_prim', 'entry_number')
        }
        
        if col_name not in editable_cols:
            return
        
        db_field, widget_type = editable_cols[col_name]
        current_value = values[col_index]
        
        # Önceki entry'yi kapat
        if self.current_edit_entry:
            self.current_edit_entry.destroy()
            self.current_edit_entry = None
        
        # Hücre koordinatlarını al
        bbox = tree.bbox(item, column)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # Widget oluştur
        if widget_type == 'combo_tur':
            self.current_edit_entry = ttk.Combobox(tree, values=TURLER, 
                                                    font=('Arial', 9), state='readonly')
            # Mevcut değeri ayarla
            current_tur = current_value
            if current_tur in TURLER:
                self.current_edit_entry.set(current_tur)
            else:
                self.current_edit_entry.set(TURLER[0])
        elif widget_type == 'combo_odeme':
            self.current_edit_entry = ttk.Combobox(tree, values=ODEME_TURLERI, 
                                                    font=('Arial', 9), state='readonly')
            # Mevcut değeri temizle (● işaretini kaldır)
            clean_value = str(current_value).replace('●', '').replace('■', '').strip()
            if 'K.KART' in clean_value or 'KART' in clean_value:
                self.current_edit_entry.set('K.KART')
            else:
                self.current_edit_entry.set('AÇIK')
        else:
            self.current_edit_entry = tk.Entry(tree, font=('Arial', 9))
            # Mevcut değeri ekle
            clean_value = str(current_value).replace(',', '') if widget_type == 'entry_number' else str(current_value)
            self.current_edit_entry.insert(0, clean_value)
            self.current_edit_entry.select_range(0, tk.END)
        
        self.current_edit_entry.place(x=x, y=y, width=width, height=height)
        self.current_edit_entry.focus_set()
        
        # Event bindings
        def save_edit(e=None):
            new_value = self.current_edit_entry.get().strip()
            
            # Brüt prim için özel işlem
            if widget_type == 'entry_number':
                try:
                    new_value = float(new_value.replace('.', '').replace(',', '.'))
                    # Komisyonu yeniden hesapla
                    tur = values[4]  # TÜR sütunu (PLAKA kaldırıldı)
                    komisyon = hesapla_komisyon(self.current_kisi, tur, new_value)
                    self.db.update_record(record_id, table_type,
                                         brut_prim=new_value,
                                         net_prim=komisyon['net_prim'],
                                         toplam_komisyon=komisyon['toplam_komisyon'],
                                         odenen_komisyon=komisyon['odenen_komisyon'])
                except:
                    self.log("⚠️ Geçersiz sayı formatı!", "warning")
                    if self.current_edit_entry:
                        self.current_edit_entry.destroy()
                        self.current_edit_entry = None
                    return
            else:
                # Normal alan güncelleme
                self.db.update_record(record_id, table_type, **{db_field: new_value})
            
            self.log(f"✅ {col_name} güncellendi", "success")
            
            if self.current_edit_entry:
                self.current_edit_entry.destroy()
                self.current_edit_entry = None
            
            self.load_data()
            self.update_summary()
        
        def cancel_edit(e=None):
            if self.current_edit_entry:
                self.current_edit_entry.destroy()
                self.current_edit_entry = None
        
        self.current_edit_entry.bind('<Return>', save_edit)
        self.current_edit_entry.bind('<Escape>', cancel_edit)
        self.current_edit_entry.bind('<FocusOut>', save_edit)
        
        if isinstance(self.current_edit_entry, ttk.Combobox):
            self.current_edit_entry.bind('<<ComboboxSelected>>', save_edit)
    
    def show_context_menu(self, event):
        """Sağ tık menüsü - AXA tablosu"""
        self.active_tree = self.tree_axa
        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()
    
    def show_context_menu_other(self, event):
        """Sağ tık menüsü - Diğer şirketler tablosu"""
        self.active_tree = self.tree_other
        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()
    
    def on_double_click_other(self, event):
        """Diğer şirketler tablosu için çift tıklama"""
        # on_double_click'i 'other' table_type ile çağır
        self.on_double_click(event, table_type='other')
    
    def delete_selected(self):
        """Seçili kayıtları sil (her iki tablodan)"""
        # Her iki tablodan seçimleri al
        selection_axa = self.tree_axa.selection()
        selection_other = self.tree_other.selection()
        
        all_selections = []
        for item in selection_axa:
            values = self.tree_axa.item(item, 'values')
            all_selections.append(('axa', values[0]))
        for item in selection_other:
            values = self.tree_other.item(item, 'values')
            all_selections.append(('other', values[0]))
        
        if not all_selections:
            messagebox.showwarning("Uyarı", "Lütfen kayıt seçin!")
            return
        
        if messagebox.askyesno("Onay", f"{len(all_selections)} kayıt silinecek. Emin misiniz?"):
            for table_type, record_id in all_selections:
                self.db.delete_record(record_id, table_type)
            
            self.log(f"🗑️ {len(all_selections)} kayıt silindi", "warning")
            self.load_data()
            self.update_summary()
    
    def delete_all_records(self):
        """Tüm veritabanını temizle ve ID'yi sıfırla - Her iki tablo"""
        # Tüm kayıtları al
        df = self.db.get_records()
        if df.empty:
            messagebox.showinfo("Bilgi", "Silinecek kayıt yok!")
            return
        
        count = len(df)
        
        if messagebox.askyesno("⚠️ DİKKAT!", 
                              f"Toplam {count} kayıt silinecek!\n\n"
                              f"TÜM VERİTABANI TEMİZLENECEK!\n"
                              f"(Hem AXA hem de diğer şirket kayıtları)\n\n"
                              f"Bu işlem geri alınamaz. Emin misiniz?"):
            # Her iki tabloyu da temizle ve ID sayaçlarını sıfırla
            try:
                with sqlite3.connect(self.db.db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM komisyonlar_axa")
                    cursor.execute("DELETE FROM komisyonlar_other")
                    cursor.execute("DELETE FROM sqlite_sequence WHERE name='komisyonlar_axa'")
                    cursor.execute("DELETE FROM sqlite_sequence WHERE name='komisyonlar_other'")
                    conn.commit()
                
                self.log(f"🗑️ Veritabanı temizlendi! {count} kayıt silindi, ID'ler sıfırlandı!", "warning")
                self.load_data()
                self.update_summary()
            except Exception as e:
                self.log(f"❌ Silme hatası: {e}", "error")
    
    def export_to_excel(self):
        """Excel'e aktar - AXA ve Diğer şirketler yan yana"""
        yil = int(self.filter_yil.get()) if self.filter_yil.get() != 'Tümü' else None
        
        # Tüm kayıtları çek
        df = self.db.get_records(kisi=self.current_kisi, yil=yil)
        
        if df.empty:
            messagebox.showwarning("Uyarı", "Dışa aktarılacak veri yok!")
            return
        
        # AXA ve diğer şirketleri ayır
        df_axa = df[df['sirket'] == 'AXA'].copy() if 'sirket' in df.columns else pd.DataFrame()
        df_other = df[df['sirket'] != 'AXA'].copy() if 'sirket' in df.columns else df.copy()
        
        # Sütun isimleri
        column_mapping = {
            'sigortali': 'SİGORTALI',
            'police_no': 'POL. NO.',
            'tarih': 'TARİH',
            'plaka': 'PLAKA',
            'tur': 'TÜR',
            'sirket': 'ŞİRKET',
            'odeme_turu': 'ÖDEME TÜRÜ',
            'brut_prim': 'BRÜT PRİM',
            'net_prim': 'NET PRİM',
            'toplam_komisyon': 'TOPLAM KOMİSYON',
            'odenen_komisyon': 'ÖDENEN KOMİSYON'
        }
        
        # Dosya adı
        yil_str = str(yil) if yil else datetime.now().strftime("%Y")
        default_name = f"{self.current_kisi}_{yil_str}.xlsx"
        
        filename = filedialog.asksaveasfilename(
            title="Excel Dosyası Kaydet",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if not filename:
            return
        
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                ws = writer.book.create_sheet(self.current_kisi)
                
                # Başlık stilleri
                header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # AXA POLİÇELERİ (Sol taraf - A sütunundan başla)
                start_col_axa = 1  # A sütunu
                
                # AXA için sütunlar (sirket olmadan)
                axa_cols = ['sigortali', 'police_no', 'tarih', 'plaka', 'tur', 'odeme_turu', 
                           'brut_prim', 'net_prim', 'toplam_komisyon', 'odenen_komisyon']
                axa_cols = [c for c in axa_cols if c in df_axa.columns]
                
                # AXA başlığı
                ws.merge_cells(start_row=1, start_column=start_col_axa, 
                              end_row=1, end_column=start_col_axa + len(axa_cols) - 1)
                cell = ws.cell(row=1, column=start_col_axa)
                cell.value = "📋 AXA POLİÇELERİ"
                cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # AXA sütun başlıkları
                for i, col in enumerate(axa_cols, start=start_col_axa):
                    cell = ws.cell(row=2, column=i)
                    cell.value = column_mapping.get(col, col.upper())
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # AXA verileri
                for row_idx, row in enumerate(df_axa[axa_cols].values, start=3):
                    for col_idx, (value, col_name) in enumerate(zip(row, axa_cols), start=start_col_axa):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.border = border
                        
                        # Sayısal sütunlar için sağa hizala ve format
                        if col_name in ['brut_prim', 'net_prim', 'toplam_komisyon', 'odenen_komisyon']:
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            if value and value != '0' and value != '-':
                                try:
                                    # Türkçe formatı koru
                                    cell.number_format = '#,##0.00'
                                except:
                                    pass
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # DİĞER ŞİRKETLER (Sağ taraf - AXA'dan sonra 2 sütun boşluk)
                start_col_other = start_col_axa + len(axa_cols) + 2
                
                # Diğer şirketler için sütunlar (sirket dahil)
                other_cols = ['sirket', 'sigortali', 'police_no', 'tarih', 'plaka', 'tur', 'odeme_turu',
                             'brut_prim', 'net_prim', 'toplam_komisyon', 'odenen_komisyon']
                other_cols = [c for c in other_cols if c in df_other.columns]
                
                # Diğer şirketler başlığı
                ws.merge_cells(start_row=1, start_column=start_col_other,
                              end_row=1, end_column=start_col_other + len(other_cols) - 1)
                cell = ws.cell(row=1, column=start_col_other)
                cell.value = "📋 DİĞER ŞİRKETLER"
                cell.fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Diğer şirketler sütun başlıkları
                for i, col in enumerate(other_cols, start=start_col_other):
                    cell = ws.cell(row=2, column=i)
                    cell.value = column_mapping.get(col, col.upper())
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # Diğer şirketler verileri
                for row_idx, row in enumerate(df_other[other_cols].values, start=3):
                    for col_idx, (value, col_name) in enumerate(zip(row, other_cols), start=start_col_other):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.border = border
                        
                        # Sayısal sütunlar için sağa hizala ve format
                        if col_name in ['brut_prim', 'net_prim', 'toplam_komisyon', 'odenen_komisyon']:
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            if value and value != '0' and value != '-':
                                try:
                                    cell.number_format = '#,##0.00'
                                except:
                                    pass
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Sütun genişliklerini optimize et
                # AXA sütunları için özel genişlikler
                axa_column_widths = {
                    'sigortali': 32,
                    'police_no': 20,
                    'tarih': 12,
                    'plaka': 11,
                    'tur': 10,
                    'odeme_turu': 13,
                    'brut_prim': 14,
                    'net_prim': 14,
                    'toplam_komisyon': 16,
                    'odenen_komisyon': 16
                }
                
                for i, col in enumerate(axa_cols, start=start_col_axa):
                    width = axa_column_widths.get(col, 15)
                    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
                
                # Diğer şirketler sütunları için özel genişlikler
                other_column_widths = {
                    'sirket': 12,
                    'sigortali': 32,
                    'police_no': 20,
                    'tarih': 12,
                    'plaka': 11,
                    'tur': 10,
                    'odeme_turu': 13,
                    'brut_prim': 14,
                    'net_prim': 14,
                    'toplam_komisyon': 16,
                    'odenen_komisyon': 16
                }
                
                for i, col in enumerate(other_cols, start=start_col_other):
                    width = other_column_widths.get(col, 15)
                    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
                
                # Satır yüksekliklerini ayarla
                ws.row_dimensions[1].height = 25
                ws.row_dimensions[2].height = 20
                
                # Varsayılan sheet'i sil
                if 'Sheet' in writer.book.sheetnames:
                    writer.book.remove(writer.book['Sheet'])
            
            self.log(f"✅ Excel kaydedildi: {os.path.basename(filename)}", "success")
            messagebox.showinfo("Başarılı", f"Excel dosyası kaydedildi:\n{filename}\n\nAXA: {len(df_axa)} kayıt\nDiğer: {len(df_other)} kayıt")
        
        except Exception as e:
            self.log(f"❌ Excel hatası: {str(e)}", "error")
            messagebox.showerror("Hata", str(e))
    
    def import_excel_data(self):
        """Excel dosyasından veri içe aktar (Rapor GUI gibi)"""
        filename = filedialog.askopenfilename(
            title="Excel Dosyası Seç",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            # Excel dosyasını oku
            df = pd.read_excel(filename)
            
            if df.empty:
                messagebox.showwarning("Uyarı", "Excel dosyası boş!")
                return
            
            # Sütun isimlerini normalize et
            df.columns = [str(c).strip().upper() for c in df.columns]
            
            # NaN değerlerini temizle
            df = df.fillna('')
            
            total_imported = 0
            total_skipped = 0
            
            for _, row in df.iterrows():
                try:
                    # Sigortalı kontrolü
                    sigortali = str(row.get('SİGORTALI', row.get('SIGORTALI', '')))
                    if not sigortali or sigortali == 'nan' or len(sigortali) < 2:
                        continue
                    
                    # Poliçe no
                    police_no = str(row.get('POL. NO.', row.get('POLİÇE NO', row.get('POLICE_NO', '')))).replace('.0', '')
                    if not police_no or police_no == 'nan':
                        continue
                    
                    # Tarih
                    tarih = row.get('TARİH', row.get('TARIH', ''))
                    if tarih:
                        if isinstance(tarih, datetime):
                            tarih = tarih.strftime('%d/%m/%Y')
                        else:
                            tarih = str(tarih)
                    else:
                        tarih = ''
                    
                    # Yıl çıkar
                    yil = datetime.now().year
                    if tarih and '/' in tarih:
                        try:
                            yil = int(tarih.split('/')[2])
                        except:
                            pass
                    
                    # Değerleri al
                    brut_prim = 0
                    for col in ['BRÜT PRİM', 'BRUT_PRIM', 'BRUT PRIM']:
                        val = row.get(col, 0)
                        if val and str(val) != 'nan':
                            try:
                                brut_prim = float(str(val).replace('.', '').replace(',', '.'))
                                break
                            except:
                                pass
                    
                    # Tür
                    tur = str(row.get('TÜR', row.get('TUR', 'TRAFİK')))
                    if tur == 'nan' or not tur:
                        tur = 'TRAFİK'
                    
                    # Komisyon hesapla
                    komisyon = hesapla_komisyon(self.current_kisi, tur, brut_prim)
                    
                    record = {
                        'kisi': self.current_kisi,
                        'ay': '-',
                        'yil': yil,
                        'sigortali': sigortali.strip(),
                        'police_no': police_no,
                        'tarih': tarih,
                        'plaka': str(row.get('PLAKA', '-')).strip() or '-',
                        'tur': tur,
                        'odeme_turu': str(row.get('ÖDEME TÜRÜ', row.get('ODEME_TURU', 'AÇIK'))).strip() or 'AÇIK',
                        'brut_prim': brut_prim,
                        'net_prim': komisyon['net_prim'],
                        'komisyon_orani': komisyon['komisyon_orani'],
                        'toplam_komisyon': komisyon['toplam_komisyon'],
                        'odeme_orani': komisyon['odeme_orani'],
                        'odenen_komisyon': komisyon['odenen_komisyon'],
                        'notlar': f'Excel import: {os.path.basename(filename)}'
                    }
                    
                    if self.db.insert_record(record):
                        total_imported += 1
                    else:
                        total_skipped += 1
                
                except Exception as row_err:
                    print(f"Satır hatası: {row_err}")
                    continue
            
            msg = f"✅ İçe aktarıldı: {total_imported}\n"
            if total_skipped > 0:
                msg += f"⚠️ Atlandı (duplicate): {total_skipped}"
            
            messagebox.showinfo("İçe Aktarma Tamamlandı", msg)
            self.log(f"✅ Excel'den {total_imported} kayıt içe aktarıldı", "success")
            self.load_data()
            self.update_summary()
        
        except Exception as e:
            self.log(f"❌ Excel okuma hatası: {str(e)}", "error")
            messagebox.showerror("Hata", str(e))
    
    def log(self, message, level="info"):
        """Log ekle"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        log_entry = f"[{timestamp}] {message}\n"
        self.log_text.insert(tk.END, log_entry, level)
        self.log_text.see(tk.END)
        self.status_bar.config(text=message)
    
    def clear_logs(self):
        """Logları temizle"""
        self.log_text.delete('1.0', tk.END)
    
    def show_about(self):
        """Hakkında"""
        about_text = """
🏢 AXA Muhasebe ve Komisyon Takip Sistemi
Versiyon: 1.0

✨ Özellikler:
• Otomatik PDF okuma ve veri çıkarma
• Kişi bazlı komisyon takibi
• Otomatik komisyon hesaplama
• Aylık Excel raporlama
• Modern kullanıcı arayüzü

📋 Komisyon Oranları:
• Trafik: %10 (tüm kişiler)
• Diğer (Yaşar/Kamil): %15
• Diğer (Tezer): %13

💰 Ödeme Oranları:
• Yaşar: %60
• Kamil/Tezer: %50
        """
        messagebox.showinfo("Hakkında", about_text)

# ==============================================================================
# ANA PROGRAM
# ==============================================================================

if __name__ == "__main__":
    root = tk.Tk()
    app = MuhasebeGUI(root)
    root.mainloop()
